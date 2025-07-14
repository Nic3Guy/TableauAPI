"""Export commands for metadata and artifacts."""

from pathlib import Path

import click
from rich.console import Console
from rich.progress import Progress

from tableauapi.core.auth import create_auth_config_from_env
from tableauapi.core.client import TableauClient
from tableauapi.core.metadata import MetadataProcessor
from tableauapi.core.storage import S3Storage, StorageManager
from tableauapi.utils.exceptions import ConfigurationError, TableauAPIError

console = Console()


@click.group(name="export")
def export_cmd():
    """Export metadata and artifacts to various formats."""
    pass


@export_cmd.command()
@click.option("--filename", "-f", help="Metadata filename to export")
@click.option(
    "--backend",
    "-b",
    type=click.Choice(["local", "s3"]),
    default="local",
    help="Storage backend to load from",
)
@click.option("--s3-bucket", help="S3 bucket name (required for S3 backend)")
@click.option("--s3-prefix", default="tableau_metadata/", help="S3 prefix")
@click.option("--output", "-o", help="Output file path")
@click.option(
    "--format",
    "-fmt",
    type=click.Choice(["json", "csv", "xlsx"]),
    default="json",
    help="Export format",
)
@click.option("--pretty", is_flag=True, help="Pretty print JSON")
def metadata(filename, backend, s3_bucket, s3_prefix, output, format, pretty):
    """Export metadata to different formats."""
    try:
        # Initialize storage
        storage_manager = StorageManager()

        if backend == "s3":
            if not s3_bucket:
                console.print("[red]S3 bucket name required for S3 backend[/red]")
                exit(1)
            s3_storage = S3Storage(s3_bucket, s3_prefix)
            storage_manager = StorageManager(s3_storage=s3_storage)

        # Load metadata
        if not filename:
            # List available files and let user choose
            if backend == "local":
                files = storage_manager.local_storage.list_metadata_files()
            else:
                files = storage_manager.s3_storage.list_metadata_files()

            if not files:
                console.print(
                    f"[yellow]No metadata files found in {backend} storage[/yellow]"
                )
                return

            console.print(
                f"[bold]Available metadata files in {backend} storage:[/bold]"
            )
            for i, file_path in enumerate(files, 1):
                console.print(f"{i}. {file_path}")

            choice = click.prompt("Select file number", type=int)
            if choice < 1 or choice > len(files):
                console.print("[red]Invalid selection[/red]")
                return

            filename = files[choice - 1]

        with console.status(f"Loading metadata from {filename}..."):
            metadata = storage_manager.load_metadata(filename, backend)

        processor = MetadataProcessor()

        # Generate output filename if not provided
        if not output:
            base_name = Path(filename).stem
            output = f"{base_name}_export.{format}"

        output_path = Path(output)

        if format == "json":
            indent = 2 if pretty else None
            json_content = processor.metadata_to_json(metadata, indent=indent)

            with open(output_path, "w", encoding="utf-8") as f:
                f.write(json_content)

            console.print(f"[green]✓ Metadata exported to {output_path}[/green]")

        elif format == "csv":
            # Export as CSV files in a directory
            output_dir = output_path.with_suffix("")
            output_dir.mkdir(parents=True, exist_ok=True)

            _export_csv(metadata, output_dir)

            console.print(f"[green]✓ Metadata exported to {output_dir}/[/green]")

        elif format == "xlsx":
            _export_xlsx(metadata, output_path)
            console.print(f"[green]✓ Metadata exported to {output_path}[/green]")

        else:
            console.print(f"[red]Unsupported format: {format}[/red]")
            exit(1)

    except Exception as e:
        console.print(f"[red]Error: {str(e)}[/red]")
        exit(1)


@export_cmd.command()
@click.option("--projects", "-p", multiple=True, help="Filter by project names")
@click.option("--owners", "-o", multiple=True, help="Filter by owner names")
@click.option("--tags", "-t", multiple=True, help="Filter by tags")
@click.option("--output", "-o", default="tableau_report.xlsx", help="Output file path")
def report(projects, owners, tags, output):
    """Generate a comprehensive report of Tableau content."""
    try:
        config = create_auth_config_from_env()
        client = TableauClient(config)
        processor = MetadataProcessor()

        with client.connection():
            with Progress() as progress:
                # Collect all data
                task = progress.add_task("Collecting data...", total=4)

                workbooks = client.get_workbooks()
                progress.update(task, advance=1)

                datasources = client.get_datasources()
                progress.update(task, advance=1)

                projects_data = client.get_projects()
                progress.update(task, advance=1)

                flows = client.get_flows()
                progress.update(task, advance=1)

                # Process data
                task = progress.add_task("Processing data...", total=4)

                processed_workbooks = [
                    processor.process_workbook(wb) for wb in workbooks
                ]
                progress.update(task, advance=1)

                processed_datasources = [
                    processor.process_datasource(ds) for ds in datasources
                ]
                progress.update(task, advance=1)

                processed_projects = [
                    processor.process_project(p) for p in projects_data
                ]
                progress.update(task, advance=1)

                processed_flows = [processor.process_flow(f) for f in flows]
                progress.update(task, advance=1)

                # Create server metadata
                server_info = client.get_server_info()
                server_metadata = processor.create_server_metadata(
                    server_info,
                    processed_workbooks,
                    processed_datasources,
                    processed_projects,
                    processed_flows,
                    config.server_url,
                    config.site_id,
                    config.site_id or "default",
                )

                # Apply filters
                if projects or owners or tags:
                    filters = {}
                    if projects:
                        filters["project_names"] = list(projects)
                    if owners:
                        filters["owner_names"] = list(owners)
                    if tags:
                        filters["tags"] = list(tags)

                    server_metadata = processor.filter_metadata(
                        server_metadata, filters
                    )

                # Export report
                task = progress.add_task("Generating report...", total=1)
                _export_xlsx(server_metadata, Path(output))
                progress.update(task, advance=1)

                console.print(f"[green]✓ Report generated: {output}[/green]")
                console.print(f"Workbooks: {len(server_metadata.workbooks)}")
                console.print(f"Data Sources: {len(server_metadata.datasources)}")
                console.print(f"Projects: {len(server_metadata.projects)}")
                console.print(f"Flows: {len(server_metadata.flows)}")

    except ConfigurationError as e:
        console.print(f"[red]Configuration error: {str(e)}[/red]")
        exit(1)
    except TableauAPIError as e:
        console.print(f"[red]API error: {str(e)}[/red]")
        exit(1)
    except Exception as e:
        console.print(f"[red]Unexpected error: {str(e)}[/red]")
        console.print_exception()
        exit(1)


def _export_csv(metadata, output_dir):
    """Export metadata as CSV files."""
    import csv

    # Export workbooks
    if metadata.workbooks:
        workbooks_file = output_dir / "workbooks.csv"
        with open(workbooks_file, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(
                [
                    "ID",
                    "Name",
                    "Description",
                    "Project",
                    "Owner",
                    "Created",
                    "Updated",
                    "Size",
                    "URL",
                    "Show Tabs",
                    "Tags",
                ]
            )

            for wb in metadata.workbooks:
                writer.writerow(
                    [
                        wb.id,
                        wb.name,
                        wb.description or "",
                        wb.project_name,
                        wb.owner_name,
                        wb.created_at.isoformat() if wb.created_at else "",
                        wb.updated_at.isoformat() if wb.updated_at else "",
                        wb.size or 0,
                        wb.webpage_url or "",
                        wb.show_tabs or False,
                        ",".join(wb.tags) if wb.tags else "",
                    ]
                )

    # Export datasources
    if metadata.datasources:
        datasources_file = output_dir / "datasources.csv"
        with open(datasources_file, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(
                [
                    "ID",
                    "Name",
                    "Description",
                    "Project",
                    "Owner",
                    "Created",
                    "Updated",
                    "Size",
                    "URL",
                    "Tags",
                ]
            )

            for ds in metadata.datasources:
                writer.writerow(
                    [
                        ds.id,
                        ds.name,
                        ds.description or "",
                        ds.project_name,
                        ds.owner_name,
                        ds.created_at.isoformat() if ds.created_at else "",
                        ds.updated_at.isoformat() if ds.updated_at else "",
                        ds.size or 0,
                        ds.webpage_url or "",
                        ",".join(ds.tags) if ds.tags else "",
                    ]
                )

    # Export projects
    if metadata.projects:
        projects_file = output_dir / "projects.csv"
        with open(projects_file, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(
                [
                    "ID",
                    "Name",
                    "Description",
                    "Parent ID",
                    "Permissions",
                    "Created",
                    "Updated",
                ]
            )

            for p in metadata.projects:
                writer.writerow(
                    [
                        p.id,
                        p.name,
                        p.description or "",
                        p.parent_id or "",
                        p.content_permissions or "",
                        p.created_at.isoformat() if p.created_at else "",
                        p.updated_at.isoformat() if p.updated_at else "",
                    ]
                )


def _export_xlsx(metadata, output_path):
    """Export metadata as Excel file."""
    try:
        import openpyxl
        from openpyxl.styles import Font

        wb = openpyxl.Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Summary sheet
        summary_sheet = wb.create_sheet("Summary")
        summary_sheet.append(["Tableau Server Metadata Summary"])
        summary_sheet.append([""])
        summary_sheet.append(["Server URL", metadata.server_url])
        summary_sheet.append(["Site Name", metadata.site_name])
        summary_sheet.append(["Generated", metadata.timestamp.isoformat()])
        summary_sheet.append([""])
        summary_sheet.append(["Content Counts"])
        summary_sheet.append(["Workbooks", len(metadata.workbooks)])
        summary_sheet.append(["Data Sources", len(metadata.datasources)])
        summary_sheet.append(["Projects", len(metadata.projects)])
        summary_sheet.append(["Flows", len(metadata.flows)])

        # Style the summary sheet
        for row in summary_sheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = Font(bold=True, size=14)

        # Workbooks sheet
        if metadata.workbooks:
            wb_sheet = wb.create_sheet("Workbooks")
            wb_sheet.append(
                [
                    "ID",
                    "Name",
                    "Description",
                    "Project",
                    "Owner",
                    "Created",
                    "Updated",
                    "Size",
                    "URL",
                    "Show Tabs",
                    "Tags",
                ]
            )

            for wb_item in metadata.workbooks:
                wb_sheet.append(
                    [
                        wb_item.id,
                        wb_item.name,
                        wb_item.description or "",
                        wb_item.project_name,
                        wb_item.owner_name,
                        wb_item.created_at.isoformat() if wb_item.created_at else "",
                        wb_item.updated_at.isoformat() if wb_item.updated_at else "",
                        wb_item.size or 0,
                        wb_item.webpage_url or "",
                        wb_item.show_tabs or False,
                        ",".join(wb_item.tags) if wb_item.tags else "",
                    ]
                )

            # Style header row
            for cell in wb_sheet[1]:
                cell.font = Font(bold=True)

        # Data Sources sheet
        if metadata.datasources:
            ds_sheet = wb.create_sheet("Data Sources")
            ds_sheet.append(
                [
                    "ID",
                    "Name",
                    "Description",
                    "Project",
                    "Owner",
                    "Created",
                    "Updated",
                    "Size",
                    "URL",
                    "Tags",
                ]
            )

            for ds in metadata.datasources:
                ds_sheet.append(
                    [
                        ds.id,
                        ds.name,
                        ds.description or "",
                        ds.project_name,
                        ds.owner_name,
                        ds.created_at.isoformat() if ds.created_at else "",
                        ds.updated_at.isoformat() if ds.updated_at else "",
                        ds.size or 0,
                        ds.webpage_url or "",
                        ",".join(ds.tags) if ds.tags else "",
                    ]
                )

            # Style header row
            for cell in ds_sheet[1]:
                cell.font = Font(bold=True)

        # Projects sheet
        if metadata.projects:
            proj_sheet = wb.create_sheet("Projects")
            proj_sheet.append(
                [
                    "ID",
                    "Name",
                    "Description",
                    "Parent ID",
                    "Permissions",
                    "Created",
                    "Updated",
                ]
            )

            for p in metadata.projects:
                proj_sheet.append(
                    [
                        p.id,
                        p.name,
                        p.description or "",
                        p.parent_id or "",
                        p.content_permissions or "",
                        p.created_at.isoformat() if p.created_at else "",
                        p.updated_at.isoformat() if p.updated_at else "",
                    ]
                )

            # Style header row
            for cell in proj_sheet[1]:
                cell.font = Font(bold=True)

        # Save workbook
        wb.save(output_path)

    except ImportError:
        console.print("[red]openpyxl library required for Excel export[/red]")
        console.print("Install with: pip install openpyxl")
        exit(1)
